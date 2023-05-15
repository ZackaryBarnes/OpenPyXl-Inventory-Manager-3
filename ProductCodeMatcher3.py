from openpyxl import Workbook, load_workbook
#-------------------------------------------------------------
#SETUP
#-------------------------------------------------------------
wb1 = load_workbook('Master Price List.xlsx')
wb2 = load_workbook('DDI Export 2.xlsx')

ws1 = wb1.active
ws2 = wb2.active
#-------------------------------------------------------------
#FUNCTIONS
#-------------------------------------------------------------
def ingestToDict(ws):
    sheet_cells = {}
    rowNum = 2
    #****row 0 is the first cell in the master "manufacturer"****
    idxValueToMatch = 3
    #****row 0 is the first cell in the master "manufacturer"****
    idxValueToCopyOver = 1

    for row in ws.iter_rows(min_row=2):
        if row[idxValueToCopyOver].value:
            if row[idxValueToMatch].value:
                valueToCopyOver = row[idxValueToCopyOver].value
                valueToMatch = row[idxValueToMatch].value
                sheet_cells[valueToMatch] = valueToCopyOver

        rowNum += 1
    return sheet_cells


def createRow(row):
    row_cells = []
    for cell in range(0, len(row)):
        row_cells.append(row[cell].value)
    return row_cells


def ingestToList(ws):
    sheet_cells = []
    for row in ws.iter_rows(min_row=2):
        newRow = createRow(row)
        sheet_cells.append(newRow)
    return sheet_cells


def mashSheets(master, template):
    matches = 0
    noMatch = 0
    duplicateMatches = 0
    nonDuplicateMatches = 0
    rowNum = 2
    #****row 0 is the first cell in the template "product(25)"****
    idxValueToMatch = 3
    
    for row in template:
        try:
            if row[idxValueToMatch] in master:
                matches += 1
                valueToCopy = master[row[idxValueToMatch]]

                #column 1 is the first cell in the template "product(25)"
                currentCellValue = ws2.cell(row=rowNum, column=6)

                if currentCellValue != valueToCopy:
                    nonDuplicateMatches += 1

                if currentCellValue == valueToCopy:
                    duplicateMatches += 1

                #column 1 is the first cell in the template "product(25)"
                ws2.cell(row=rowNum, column=6, value=valueToCopy)
            else:
                noMatch += 1
        except Exception as e:
            print(str(e))

        rowNum += 1
            
    print(f"Matches: {matches}, No match: {noMatch}, Duplicate Matches: {duplicateMatches}, Non-duplicate matches: {nonDuplicateMatches}")
#-------------------------------------------------------------
#IMPLEMENTATION
#-------------------------------------------------------------
templateList = ingestToList(ws2)
masterDict = ingestToDict(ws1)
mashSheets(masterDict, templateList)
#-------------------------------------------------------------
#CLOSING
#-------------------------------------------------------------
wb1.save('Master Price List.xlsx')
wb2.save('DDI Export 2.xlsx')